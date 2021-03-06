from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .forms import *
from django.contrib.auth import authenticate, logout, login
from .decorators import *
from django.contrib import messages
from .models import *
from django.shortcuts import get_object_or_404, redirect, render
import csv
import xlwt
import datetime
from django.http import JsonResponse
import sqlite3



@login_required(login_url='loly_login')
def exp(request):

    labsvleid=[
        'NMRL',
        'Mpilo',
        'Mutare',
        'BRIDH',
        'Gweru',
        'Chinhoyi',
        'Masvingo',
        'Victoria Falls', 
        'Bindura',
        'Kadoma',
        'Marondera',
        'St Lukes',
        'Gwanda']

    populateVleidSheets(labsvleid)


    from openpyxl.writer.excel import save_virtual_workbook
    workbook = load_workbook(filename="hello_eid.xlsx")
    data =save_virtual_workbook(workbook)



    response=HttpResponse(content=data,content_type='application/ms-excel')
    response['Content-Disposition']='attachment; filename=eid' + '.xlsx'




    # workbook = load_workbook(filename="C:\\Users\\kaste\\OneDrive\\Desktop\\meals postgres\\meal\\pages\\covid.xlsx")
    # #lst = workbook.sheetnames
    # ws = workbook[workbook.sheetnames[0]]
    # sheet = workbook.active


    # getRecord(getReportingWeek(),request.user.lab,'covid','specimenrun',workbook=workbook ,ws= sheet,lst=getCovidSpecimenRunList())

    # getRecord(getReportingWeek(),request.user.lab,'covid','specimenreceived',workbook=workbook ,ws= sheet,lst=getCovidSpecimenReceiveList())
    # getRecord(getReportingWeek(),request.user.lab,'covid','machine',workbook=workbook ,ws= sheet,lst=getCovidMachineList())
    # getRecord(getReportingWeek(),request.user.lab,'covid','general1',workbook=workbook ,ws= sheet,lst=getCovidGeneralList())
    # getRecord(getReportingWeek(),request.user.lab,'covid','general2',workbook=workbook ,ws= sheet,lst=getCovidGeneralList())
    # getRecord(getReportingWeek(),request.user.lab,'covid','general3',workbook=workbook ,ws= sheet,lst=getCovidGeneralList())

    return response





















    # context = {}
    # columns=[]
    # record=[]

    # try:
    #     sqliteConnection = sqlite3.connect('db.sqlite3')
    #     cursorC = sqliteConnection.cursor()
    #     cursorR = sqliteConnection.cursor()
    #     print("Database success")


    #     f = open("tables.txt",'r')
    #     tables=f.readlines()
    #     i=0
    #     for table in tables:
    #         # sql_select_Query = "select name from sqlite_master where type='table' and name like 'pages_%';"
    #         sql_select_Query = "PRAGMA table_info( "+table.strip()+" );"
    #         cursorC.execute(sql_select_Query)
    #         columns.append(cursorC.fetchall())
    #         #print(columns)

    #         sql_select_Query1 = "select * from "+table+";"
    #         cursorR.execute(sql_select_Query1)
    #         record.append(cursorR.fetchall())
    #         #print(record)

    #         i=i+1

    #         context={'columns': columns,
    #                 'record':record
    #         }

    #         # f=open('columns.txt', 'w')
            
    #         # for table in record:
    #         #     f.write(str(table) + '\n')

    #         # f.close()

    # except sqlite3.Error as error:
    #     print("ndatadza sha", error)

    # # finally:
    # #     if sqliteConnection:
    # #         sqliteConnection.close()
    # #         print("ndavhara")
    # #     cursor.close()




def index(request):
    return HttpResponse("Hello, world. You're at the pages index.")



def home(request):

    context={'user':request.user}
    print(context['user'], 'fghjk')




    return render(request, 'landing.html', context)




@unauthenticated_user
def loly_login(request):
    user = request.user
    if request.method == 'POST':
        username = request.POST.get('username')
        password =request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        print(user)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.info(request, 'Username OR password is incorrect')
    
    context = {'user':user}
    return render(request, 'registration/login.html', context)

# @allowed_users
# def special(request):
#     return HttpResponse("You are logged in !")

def logoutUser(request):
	logout(request)
	return redirect('loly_login')


#function to get current reporting week
def getReportingWeek():
    #do calcs

    #for every new reporting week we nned to create new empty total records for that week to be used that week
    #ie for every lab total create an total for that lab to be updated by incomming records

    #     labchoices=[
    #     ('NMRL','NMRL'),
    #     ('Mpilo','Mpilo'),
    #     ('BRIDH','BRIDH'),
    #     ('NTBRL','NTBRL'),
    #     ('Gweru','Gweru'),
    #     ('Chinhoyi','Chinhoyi'),
    #     ('Masvingo','Masvingo'),
    #     ('eid','eid'),
    #     ('Victoria Falls', 'Victoria Falls'),
    #     ('Bindura','Bindura'),
    #     ('Kadoma','Kadoma'),
    #     ('Marondera','Marondera'),
    #     ('St Lukes', 'St Lukes'),
    #     ('Gwanda','Gwanda'),
        
    # ]

    return "1"


















#------------------------------lab vl--------------------------------------------------------------------------------------------------------












@login_required(login_url='loly_login')
def labvltop(request):
    if request.method == 'POST':
        form = SpecimensrunbrtivlweeklyForm(request.POST)
        if form.is_valid():
            form.save()
            #return render(request, 'success.html')
    form = SpecimensrunbrtivlweeklyForm()
    context = {'form': form , 'vl':True,'eid':False,'covid':False}
    return render(request, 'masving_brti_vl_weekly_statistics_tool_31-6_june_2021/top.html', context)




@login_required(login_url='loly_login')
def labvlrun(request):
    lv=True
    covid=False
    eid=False
    if request.method == 'POST':
        form = SpecimensrunbrtivlweeklyForm(request.POST)
        if form.is_valid():
            
            updateTotal = specimensrunbrtivleid.objects.get(key="vl", lab=request.user.lab,reportingweek=getReportingWeek() )

            updateTotal.testsdonerochenumberofsamplesreceivedthisweekplasma = updateTotal.testsdonerochenumberofsamplesreceivedthisweekplasma + form.cleaned_data.get('testsdonerochenumberofsamplesreceivedthisweekplasma')
            updateTotal.testsdonerochenumberofsamplescarriedoverpreviousweeksplasma = updateTotal.testsdonerochenumberofsamplescarriedoverpreviousweeksplasma + form.cleaned_data.get('testsdonerochenumberofsamplescarriedoverpreviousweeksplasma')
            
            updateTotal.testsdonerochefailedbuteligibaleforrepeatplasma += form.cleaned_data.get('testsdonerochefailedbuteligibaleforrepeatplasma')
            updateTotal.testsdonerochefailedbutnoteligibaleforrepeatplasma += form.cleaned_data.get('testsdonerochefailedbutnoteligibaleforrepeatplasma')
            updateTotal.testsdonerocherepeatplasma += form.cleaned_data.get('testsdonerocherepeatplasma')
            updateTotal.testsdonerochefailedafterrepeattestingplasma += form.cleaned_data.get('testsdonerochefailedafterrepeattestingplasma')
            updateTotal.testsdonerochenumberofsamplesreceivedthisweekdbs += form.cleaned_data.get('testsdonerochenumberofsamplesreceivedthisweekdbs')

            updateTotal.testsdonerochenumberofsamplescarriedoverpreviousweeksdbs += form.cleaned_data.get('testsdonerochenumberofsamplescarriedoverpreviousweeksdbs')
            updateTotal.testsdonerochefailedbuteligibaleforrepeatdbs += form.cleaned_data.get('testsdonerochefailedbuteligibaleforrepeatdbs')
            updateTotal.testsdonerochefailedbutnoteligibaleforrepeatdbs += form.cleaned_data.get('testsdonerochefailedbutnoteligibaleforrepeatdbs')
            updateTotal.testsdonerocherepeatdbs += form.cleaned_data.get('testsdonerocherepeatdbs')
            updateTotal.testsdonerochefailedafterrepeattestingdbs += form.cleaned_data.get('testsdonerochefailedafterrepeattestingdbs')
            updateTotal.testsdonebmxnumberofsamplesreceivedthisweekplasma += form.cleaned_data.get('testsdonebmxnumberofsamplesreceivedthisweekplasma')

            updateTotal.testsdonebmxnumberofsamplescarriedoverpreviousweeksplasma += form.cleaned_data.get('testsdonebmxnumberofsamplescarriedoverpreviousweeksplasma')
            updateTotal.testsdonebmxfailedbuteligibaleforrepeatplasma += form.cleaned_data.get('testsdonebmxfailedbuteligibaleforrepeatplasma')
            updateTotal.testsdonebmxfailedbutnoteligibaleforrepeatplasma += form.cleaned_data.get('testsdonebmxfailedbutnoteligibaleforrepeatplasma')
            updateTotal.testsdonebmxrepeatplasma += form.cleaned_data.get('testsdonebmxrepeatplasma')
            updateTotal.testsdonebmxfailedafterrepeattestingplasma += form.cleaned_data.get('testsdonebmxfailedafterrepeattestingplasma')
            updateTotal.testsdonebmxnumberofsamplesreceivedthisweekdbs += form.cleaned_data.get('testsdonebmxnumberofsamplesreceivedthisweekdbs')

            updateTotal.testsdonebmxnumberofsamplescarriedoverpreviousweeksdbs += form.cleaned_data.get('testsdonebmxnumberofsamplescarriedoverpreviousweeksdbs')
            updateTotal.testsdonebmxfailedbuteligibaleforrepeatdbs += form.cleaned_data.get('testsdonebmxfailedbuteligibaleforrepeatdbs')
            updateTotal.testsdonebmxfailedbutnoteligibaleforrepeatdbs += form.cleaned_data.get('testsdonebmxfailedbutnoteligibaleforrepeatdbs')
            updateTotal.testsdonebmxrepeatdbs += form.cleaned_data.get('testsdonebmxrepeatdbs')
            updateTotal.testsdonebmxfailedafterrepeattestingdbs += form.cleaned_data.get('testsdonebmxfailedafterrepeattestingdbs')
            updateTotal.testsdoneabbottnumberofsamplesreceivedthisweekplasma += form.cleaned_data.get('testsdoneabbottnumberofsamplesreceivedthisweekplasma')

            updateTotal.testsdoneabbottnumberofsamplescarriedoverpreviousweeksplasma += form.cleaned_data.get('testsdoneabbottnumberofsamplescarriedoverpreviousweeksplasma')
            updateTotal.testsdoneabbottfailedbuteligibaleforrepeatplasma += form.cleaned_data.get('testsdoneabbottfailedbuteligibaleforrepeatplasma')
            updateTotal.testsdoneabbottfailedbutnoteligibaleforrepeatplasma += form.cleaned_data.get('testsdoneabbottfailedbutnoteligibaleforrepeatplasma')
            updateTotal.testsdoneabbottrepeatplasma += form.cleaned_data.get('testsdoneabbottrepeatplasma')
            updateTotal.testsdoneabbottfailedafterrepeattestingplasma += form.cleaned_data.get('testsdoneabbottfailedafterrepeattestingplasma')
            updateTotal.testsdoneabbottnumberofsamplesreceivedthisweekdbs += form.cleaned_data.get('testsdoneabbottnumberofsamplesreceivedthisweekdbs')

            updateTotal.testsdoneabbottnumberofsamplescarriedoverpreviousweeksdbs += form.cleaned_data.get('testsdoneabbottnumberofsamplescarriedoverpreviousweeksdbs')
            updateTotal.testsdoneabbottfailedbuteligibaleforrepeatdbs += form.cleaned_data.get('testsdoneabbottfailedbuteligibaleforrepeatdbs')
            updateTotal.testsdoneabbottfailedbutnoteligibaleforrepeatdbs += form.cleaned_data.get('testsdoneabbottfailedbutnoteligibaleforrepeatdbs')
            updateTotal.testsdoneabbottrepeatdbs += form.cleaned_data.get('testsdoneabbottrepeatdbs')
            updateTotal.testsdoneabbottfailedafterrepeattestingdbs += form.cleaned_data.get('testsdoneabbottfailedafterrepeattestingdbs')
            updateTotal.testsdonehologicpanthernumberofsamplesreceivedthisweekplasma += form.cleaned_data.get('testsdonehologicpanthernumberofsamplesreceivedthisweekplasma')

            updateTotal.tests2 += form.cleaned_data.get('tests2')
            updateTotal.testsdonehologicpantherfailedbuteligibaleforrepeatplasma += form.cleaned_data.get('testsdonehologicpantherfailedbuteligibaleforrepeatplasma')
            updateTotal.testsdonehologicpantherfailedbutnoteligibaleforrepeatplasma += form.cleaned_data.get('testsdonehologicpantherfailedbutnoteligibaleforrepeatplasma')
            updateTotal.testsdonehologicpantherrepeatplasma += form.cleaned_data.get('testsdonehologicpantherrepeatplasma')
            updateTotal.testsdonehologicpantherfailedafterrepeattestingplasma += form.cleaned_data.get('testsdonehologicpantherfailedafterrepeattestingplasma')
            updateTotal.testsdonehologicpanthernumberofsamplesreceivedthisweekdbs += form.cleaned_data.get('testsdonehologicpanthernumberofsamplesreceivedthisweekdbs')

            updateTotal.tests1 += form.cleaned_data.get('tests1')
            updateTotal.testsdonehologicpantherfailedbuteligibaleforrepeatdbs += form.cleaned_data.get('testsdonehologicpantherfailedbuteligibaleforrepeatdbs')
            updateTotal.testsdonehologicpantherfailedbutnoteligibaleforrepeatdbs += form.cleaned_data.get('testsdonehologicpantherfailedbutnoteligibaleforrepeatdbs')
            updateTotal.testsdonehologicpantherrepeatdbs += form.cleaned_data.get('testsdonehologicpantherrepeatdbs')
            updateTotal.testsdonehologicpantherfailedafterrepeattestingdbs += form.cleaned_data.get('testsdonehologicpantherfailedafterrepeattestingdbs')
            updateTotal.totaltestsdone += form.cleaned_data.get('totaltestsdone')

            updateTotal.totalrepeats += form.cleaned_data.get('totalrepeats')
            updateTotal.totalpatientsrun += form.cleaned_data.get('totalpatientsrun')
            updateTotal.targetsweekly += form.cleaned_data.get('targetsweekly')
            updateTotal.percentagetargetsachievements += form.cleaned_data.get('percentagetargetsachievements')
            updateTotal.percentageerrorraterocheplasma += form.cleaned_data.get('percentageerrorraterocheplasma')
            updateTotal.percentageerrorraterochedbs += form.cleaned_data.get('percentageerrorraterochedbs')

            updateTotal.percentageerrorratebmxplasma += form.cleaned_data.get('percentageerrorratebmxplasma')
            updateTotal.percentageerrorratebmxdbs += form.cleaned_data.get('percentageerrorratebmxdbs')
            updateTotal.percentageerrorrateabbottplasma += form.cleaned_data.get('percentageerrorrateabbottplasma')
            updateTotal.percentageerrorrateabbottdbs += form.cleaned_data.get('percentageerrorrateabbottdbs')
            updateTotal.percentageerrorratehologicpantherplasma += form.cleaned_data.get('percentageerrorratehologicpantherplasma')
            updateTotal.percentageerrorratehologicpantherdbs += form.cleaned_data.get('percentageerrorratehologicpantherdbs')

            updateTotal.totalncsfromaudit += form.cleaned_data.get('totalncsfromaudit')
            updateTotal.ncsnotyetclosed += form.cleaned_data.get('ncsnotyetclosed')
            updateTotal.lab = request.user.lab
            updateTotal.key = "vl"
            updateTotal.ncsclosedthisweek += form.cleaned_data.get('ncsclosedthisweek')

            updateTotal.save()

            new_form = form.save(commit=False)

            new_form.lab =request.user.lab
            new_form.reportingweek = getReportingWeek()


            new_form.user=request.user
            new_form.save()


            #return render(request, 'success.html')
        else:
            print(form.errors.as_data())
            #return render(request, 'success.html')
    form = SpecimensrunbrtivlweeklyForm()
    
    context = {'form': form, 'vl':True,'eid':False,'covid':False}
    return render(request, 'masving_brti_vl_weekly_statistics_tool_31-6_june_2021/Specimensrun.html', context)

@login_required(login_url='loly_login')
def labvlrecieved(request):
    if request.method == 'POST':
        form = SpecimensrecievedbrtivlweeklyForm(request.POST)
        if form.is_valid():
            updateTotal = specimensreceivedbrtivleid.objects.get(key="vl", lab=request.user.lab,reportingweek=getReportingWeek() )

            updateTotal.samplescarriedoverpreviousweeksnevertestedplasma += form.cleaned_data.get('samplescarriedoverpreviousweeksnevertestedplasma')
            updateTotal.samplescarriedoverpreviousweeknevertesteddbs += form.cleaned_data.get('samplescarriedoverpreviousweeknevertesteddbs')

            updateTotal.samplescarriedoverpreviousfailedsamplesplasma += form.cleaned_data.get('samplescarriedoverpreviousfailedsamplesplasma')
            updateTotal.samplescarriedoverpreviousfailedsamplesdbs += form.cleaned_data.get('samplescarriedoverpreviousfailedsamplesdbs')


            updateTotal.samplesreceivedcurrentweekplasma += form.cleaned_data.get('samplesreceivedcurrentweekplasma')
            updateTotal.samplesreceivedcurrentweekdbs += form.cleaned_data.get('samplesreceivedcurrentweekdbs')

            updateTotal.samplesrejectedcurrentweekplasma += form.cleaned_data.get('samplesrejectedcurrentweekplasma')
            updateTotal.samplesrejectedcurrentweekdbs += form.cleaned_data.get('samplesrejectedcurrentweekdbs')

            updateTotal.totalsamplesreceivedcurrentweekplasma += form.cleaned_data.get('totalsamplesreceivedcurrentweekplasma')
            updateTotal.totalsamplesreceivedcurrentweekdbs += form.cleaned_data.get('totalsamplesreceivedcurrentweekdbs')

            updateTotal.numberofsamplesenteredintolimsondayofarrivalplasma += form.cleaned_data.get('numberofsamplesenteredintolimsondayofarrivalplasma')
            updateTotal.numberofsamplesenteredintolimsondayofarrivaldbs += form.cleaned_data.get('numberofsamplesenteredintolimsondayofarrivaldbs')

            updateTotal.numberofsamplesenteredintolimsafterdayofarrivalplasma += form.cleaned_data.get('numberofsamplesenteredintolimsafterdayofarrivalplasma')
            updateTotal.numberofsamplesenteredintolimsafterdayofarrivaldbs += form.cleaned_data.get('numberofsamplesenteredintolimsafterdayofarrivaldbs')

            updateTotal.numberofhourslimswasfunctional += form.cleaned_data.get('numberofhourslimswasfunctional')
            updateTotal.totalsamplescurrentandcarryoverplasma += form.cleaned_data.get('totalsamplescurrentandcarryoverplasma')

            updateTotal.totalsamplescurrentandcarryoverdbs += form.cleaned_data.get('totalsamplescurrentandcarryoverdbs')
            updateTotal.samplesrefferedplasma += form.cleaned_data.get('samplesrefferedplasma')

            updateTotal.samplesreffereddbs += form.cleaned_data.get('samplesreffereddbs')
            updateTotal.labsamplesrefferedto += form.cleaned_data.get('labsamplesrefferedto')

            updateTotal.percentagerejectionrateplasma += form.cleaned_data.get('percentagerejectionrateplasma')
            updateTotal.percentagerejectionratedbs += form.cleaned_data.get('percentagerejectionratedbs')

            updateTotal.numberofresultsprintedfromlimsplasma += form.cleaned_data.get('numberofresultsprintedfromlimsplasma')
            updateTotal.numberofresultsprintedfromlimsdbs += form.cleaned_data.get('numberofresultsprintedfromlimsdbs')

            updateTotal.totalresultsdispatchedbylabplasma += form.cleaned_data.get('totalresultsdispatchedbylabplasma')
            updateTotal.totalresultsdispatchedbylabdbs += form.cleaned_data.get('totalresultsdispatchedbylabdbs')

            updateTotal.totalresultsdispatchedbylabviasmsplasma += form.cleaned_data.get('totalresultsdispatchedbylabviasmsplasma')
            updateTotal.totalresultsdispatchedbylabviasmsdbs += form.cleaned_data.get('totalresultsdispatchedbylabviasmsdbs')

            updateTotal.reasonsforrejectionssamplequalitycompromisedplasma += form.cleaned_data.get('reasonsforrejectionssamplequalitycompromisedplasma')
            updateTotal.reasonsforrejectionssamplequalitycompromiseddbs += form.cleaned_data.get('reasonsforrejectionssamplequalitycompromiseddbs')

            updateTotal.reasons6 += form.cleaned_data.get('reasons6')
            updateTotal.reasons5 += form.cleaned_data.get('reasons5')

            updateTotal.reasons2 += form.cleaned_data.get('reasons2')
            updateTotal.reasons1 += form.cleaned_data.get('reasons1')

            updateTotal.reasons4 += form.cleaned_data.get('reasons4')
            updateTotal.reason3 += form.cleaned_data.get('reasons3')

            updateTotal.reasonsforrejectionssamplequalitycompromisedqdacheckplasma += form.cleaned_data.get('reasonsforrejectionssamplequalitycompromisedqdacheckplasma')
            updateTotal.reasonsforrejectionssamplequalitycompromisedqdacheckdbs += form.cleaned_data.get('reasonsforrejectionssamplequalitycompromisedqdacheckdbs')

            updateTotal.reasonsforsamplerefferalreagentorkitstockoutplasma += form.cleaned_data.get('reasonsforsamplerefferalreagentorkitstockoutplasma')
            updateTotal.reasonsforsamplerefferalreagentorkitstockoutdbs += form.cleaned_data.get('reasonsforsamplerefferalreagentorkitstockoutdbs')

            updateTotal.reasonsforsamplerefferalinstrumentmechanicalfailureplasma += form.cleaned_data.get('reasonsforsamplerefferalinstrumentmechanicalfailureplasma')
            updateTotal.reasonsforsamplerefferalinstrumentmechanicalfailuredbs += form.cleaned_data.get('reasonsforsamplerefferalinstrumentmechanicalfailuredbs')

            updateTotal.reasonsforsamplerefferalinsufficientinstrumentcapacityplasma += form.cleaned_data.get('reasonsforsamplerefferalinsufficientinstrumentcapacityplasma')
            updateTotal.reasonsforsamplerefferalinsufficientinstrumentcapacitydbs += form.cleaned_data.get('reasonsforsamplerefferalinsufficientinstrumentcapacitydbs')

            updateTotal.reasonsforsamplerefferalinsufficienthrcapacityplasma += form.cleaned_data.get('reasonsforsamplerefferalinsufficienthrcapacityplasma')
            updateTotal.reasonsforsamplerefferalinsufficienthrcapacitydbs += form.cleaned_data.get('reasonsforsamplerefferalinsufficienthrcapacitydbs')

            updateTotal.reasonsforsamplerefferaldqacheckplasma += form.cleaned_data.get('reasonsforsamplerefferaldqacheckplasma')
            updateTotal.reasonsforsamplerefferaldqacheckdbs += form.cleaned_data.get('reasonsforsamplerefferaldqacheckdbs')

            updateTotal.comments += form.cleaned_data.get('comments')

            updateTotal.save()

            new_form = form.save(commit=False)

            new_form.lab =request.user.lab
            new_form.reportingweek = getReportingWeek()


            new_form.user=request.user
            new_form.save()


            #return render(request, 'success.html')
        else:
            print(form.errors.as_data())
            #return render(request, 'success.html')
    form = SpecimensrecievedbrtivlweeklyForm()
    context = {'form': form , 'vl':True,'eid':False,'covid':False}
    return render(request, 'masving_brti_vl_weekly_statistics_tool_31-6_june_2021/specimenreceived.html', context)

@login_required(login_url='loly_login')
def labvlfailure(request):
    if request.method == 'POST':
        form = ReasonsforfailurebrtivlweeklyForm(request.POST)
        if form.is_valid():
            updateTotal = reasonsforfailurebrtivleid.objects.get(key="vl", lab=request.user.lab,reportingweek=getReportingWeek() )

            updateTotal.rocheplasmanumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetosamplequalityissues')
            updateTotal.rocheplasmanumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetoreagentqualityissues')
            updateTotal.rocheplasmanumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetoduetoqcfailure')
            updateTotal.rocheplasmanumberoffailedtestsduetopowerfailure += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetopowerfailure')
            updateTotal.rocheplasmanumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetomechanicalfailure')
            updateTotal.rocheplasmanumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetoprocessingerror')
            updateTotal.rocheplasmanumberoffailedtestsduetoother += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetoother')
            updateTotal.rochedqacheckplasma += form.cleaned_data.get('rochedqacheckplasma')
        

            updateTotal.rochedbsnumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('rochedbsnumberoffailedtestsduetosamplequalityissues')
            updateTotal.rochedbsnumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('rochedbsnumberoffailedtestsduetoreagentqualityissues')
            updateTotal.rochedbsnumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('rochedbsnumberoffailedtestsduetoduetoqcfailure')
            updateTotal.rochedbsnumberoffailedtestsduetopowerfailure += form.cleaned_data.get('rochedbsnumberoffailedtestsduetopowerfailure')
            updateTotal.rochedbsnumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('rochedbsnumberoffailedtestsduetomechanicalfailure')
            updateTotal.rochedbsnumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('rochedbsnumberoffailedtestsduetoprocessingerror')
            updateTotal.rochedbsnumberoffailedtestsduetoother += form.cleaned_data.get('rochedbsnumberoffailedtestsduetoother')
            updateTotal.rochedqacheckdbs += form.cleaned_data.get('rochedqacheckdbs')


            updateTotal.bmxplasmanumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetosamplequalityissues')
            updateTotal.bmxplasmanumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetoreagentqualityissues')
            updateTotal.bmxplasmanumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetoduetoqcfailure')
            updateTotal.bmxplasmanumberoffailedtestsduetopowerfailure += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetopowerfailure')


            updateTotal.bmxplasmanumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetomechanicalfailure')
            updateTotal.bmxplasmanumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetoprocessingerror')
            updateTotal.bmxplasmanumberoffailedtestsduetoother += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetoother')
            updateTotal.bmxdqacheckplasma += form.cleaned_data.get('bmxdqacheckplasma')


            updateTotal.bmxdbsnumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetosamplequalityissues')
            updateTotal.bmxdbsnumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetoreagentqualityissues')
            updateTotal.bmxdbsnumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetoduetoqcfailure')
            updateTotal.bmxdbsnumberoffailedtestsduetopowerfailure += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetopowerfailure')
            updateTotal.bmxdbsnumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetomechanicalfailure')
            updateTotal.bmxdbsnumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetoprocessingerror')




            updateTotal.bmxdbsnumberoffailedtestsduetoother += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetoother')
            updateTotal.bmxdqacheckdbs += form.cleaned_data.get('bmxdqacheckdbs')
            updateTotal.abbottplasmanumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetosamplequalityissues')
            updateTotal.abbottplasmanumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetoreagentqualityissues')
            updateTotal.abbottplasmanumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetoduetoqcfailure')
            updateTotal.abbottplasmanumberoffailedtestsduetopowerfailure += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetopowerfailure')
            updateTotal.abbottplasmanumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetomechanicalfailure')
            updateTotal.abbottplasmanumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetoprocessingerror')
            updateTotal.abbottplasmanumberoffailedtestsduetoother += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetoother')
            updateTotal.abbottdqacheckplasma += form.cleaned_data.get('abbottdqacheckplasma')
            updateTotal.abbottdbsnumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetosamplequalityissues')
            updateTotal.abbottdbsnumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetoreagentqualityissues')
            updateTotal.abbottdbsnumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetoduetoqcfailure')
            updateTotal.abbottdbsnumberoffailedtestsduetopowerfailure += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetopowerfailure')



            updateTotal.abbottdbsnumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetomechanicalfailure')
            updateTotal.abbottdbsnumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetoprocessingerror')
            updateTotal.abbottdbsnumberoffailedtestsduetoother += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetoother')
            updateTotal.abbottdqacheckdbs += form.cleaned_data.get('abbottdqacheckdbs')


            updateTotal.hologicpantherplasmanumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('hologicpantherplasmanumberoffailedtestsduetosamplequalityissues')
            updateTotal.hpantherplasmanumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('hpantherplasmanumberoffailedtestsduetoreagentqualityissues')
            updateTotal.hologicpantherplasmanumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('hologicpantherplasmanumberoffailedtestsduetoduetoqcfailure')
            updateTotal.hologicpantherplasmanumberoffailedtestsduetopowerfailure += form.cleaned_data.get('hologicpantherplasmanumberoffailedtestsduetopowerfailure')
            updateTotal.hologicpantherplasmanumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('hologicpantherplasmanumberoffailedtestsduetomechanicalfailure')
            updateTotal.hologicpantherplasmanumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('hologicpantherplasmanumberoffailedtestsduetoprocessingerror')
            updateTotal.hologicpantherplasmanumberoffailedtestsduetoother += form.cleaned_data.get('hologicpantherplasmanumberoffailedtestsduetoother')
            updateTotal.hologicpantherdqacheckplasma += form.cleaned_data.get('hologicpantherdqacheckplasma')
            updateTotal.hologicpantherdbsnumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetosamplequalityissues')
            updateTotal.hologicpantherdbsnumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetoreagentqualityissues')


            updateTotal.hologicpantherdbsnumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetoduetoqcfailure')
            updateTotal.hologicpantherdbsnumberoffailedtestsduetopowerfailure += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetopowerfailure')
            updateTotal.hologicpantherdbsnumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetomechanicalfailure')
            updateTotal.hologicpantherdbsnumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetoprocessingerror')
            updateTotal.hologicpantherdbsnumberoffailedtestsduetoother += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetoother')
            updateTotal.hologicpantherdqacheckdbs += form.cleaned_data.get('hologicpantherdqacheckdbs')


            updateTotal.save()

            new_form = form.save(commit=False)

            new_form.lab =request.user.lab
            new_form.reportingweek = getReportingWeek()


            new_form.user=request.user
            new_form.save()


            #return render(request, 'success.html')
        else:
            print(form.errors.as_data())
            #return render(request, 'success.html')
    form = ReasonsforfailurebrtivlweeklyForm()
    context = {'form': form , 'vl':True,'eid':False,'covid':False}
    return render(request, 'masving_brti_vl_weekly_statistics_tool_31-6_june_2021/Reasonsforfailures.html', context)

@login_required(login_url='loly_login')
def labvlelectric(request):
    if request.method == 'POST':
        form = ElectricoutagebrtivlweeklyForm(request.POST)
        if form.is_valid():
            updateTotal = Electricoutagebrtivleid.objects.get(key="vl", lab=request.user.lab,reportingweek=getReportingWeek() )

            updateTotal.reportingweek += form.cleaned_data.get('reportingweek')
            updateTotal.numberofhourswithnoelectricityperday += form.cleaned_data.get('numberofhourswithnoelectricityperday')
            updateTotal.numberofhoursgeneratorwasonperday += form.cleaned_data.get('numberofhoursgeneratorwasonperday')
            updateTotal.litresoffueladdedtogeneratorperday += form.cleaned_data.get('litresoffueladdedtogeneratorperday')


            updateTotal.numberofhoursmachineswasnotbeingusedduetopowercutperday += form.cleaned_data.get('numberofhoursmachineswasnotbeingusedduetopowercutperday')
            updateTotal.totaltestsdoneperdayusinggenerator += form.cleaned_data.get('totaltestsdoneperdayusinggenerator')


            updateTotal.save()

            new_form = form.save(commit=False)

            new_form.lab =request.user.lab
            new_form.reportingweek = getReportingWeek()


            new_form.user=request.user
            new_form.save()


            #return render(request, 'success.html')
        else:
            print(form.errors.as_data())
            #return render(request, 'success.html')
    form = ElectricoutagebrtivlweeklyForm()
    context = {'form': form , 'vl':True,'eid':False,'covid':False}
    return render(request, 'masving_brti_vl_weekly_statistics_tool_31-6_june_2021/electricoutagestool.html', context)



#-------------------------lab cov 19-------------------------------------------------------------------------------------------------------------------


#doesnt exist well it does but we removed it


@login_required(login_url='loly_login')
def labcov19top(request):

    #before save create or update the unique row called total to be sent over as an entry to brti cov 19 table
    if request.method == 'POST':
        form = SpecimensrunbrtivlweeklyForm(request.POST)
        if form.is_valid():
            brticov19SpecimenRunrecord = covRecordsSpecimenRun()




            form.save()
            #return render(request, 'success.html')
    form = SpecimensrunbrtivlweeklyForm()
    context = {'form': form , 'vl':False,'eid':False,'covid':True}
    return render(request, 'masvingo_brti_covid_19_weekly_statistics_tool_31-6_June_2021/top.html', context)

@login_required(login_url='loly_login')
def labcov19run(request):
        #before save create or update the unique row called total to be sent over as an entry to brti cov 19 table
    # updateTotal = Specimensruncovid19.objects.get(dayofweek="Total")
    # for row in updateTotal.__dict__:
    #     print(row)

    form = Specimensruncovid19Form()
    if request.method == 'POST':
        form = Specimensruncovid19Form(request.POST)
        if form.is_valid():
            print('hello there mate')
            #get the object for brti cov19 and disect the form object and assign or update the total entry and save
            updateTotal = specimensrunbrticovid19.objects.get(dayofweek="Total", lab=request.user.lab,reportingweek=getReportingWeek() )
            

            
            updateTotal.testsdoneabbottrun +=   form.cleaned_data.get('testsdoneabbottrun')





            updateTotal.testsdoneabbottfailedbuteligibaleforrepeat = updateTotal.testsdoneabbottfailedbuteligibaleforrepeat + form.cleaned_data.get('testsdoneabbottfailedbuteligibaleforrepeat')
            updateTotal.testsdoneabbottfailedbutnoteligibaleforrepeat = updateTotal.testsdoneabbottfailedbutnoteligibaleforrepeat + form.cleaned_data.get('testsdoneabbottfailedbutnoteligibaleforrepeat')
            updateTotal.testsdoneabbottrepeat = updateTotal.testsdoneabbottrepeat + form.cleaned_data.get('testsdoneabbottrepeat')


            updateTotal.testsdonebmxrun = updateTotal.testsdonebmxrun + form.cleaned_data.get('testsdonebmxrun')
            updateTotal.testsdonebmxfailedbuteligibaleforrepeat = updateTotal.testsdonebmxfailedbuteligibaleforrepeat + form.cleaned_data.get('testsdonebmxfailedbuteligibaleforrepeat')
            updateTotal.testsdonebmxfailedbutnoteligibaleforrepeat = updateTotal.testsdonebmxfailedbutnoteligibaleforrepeat + form.cleaned_data.get('testsdonebmxfailedbutnoteligibaleforrepeat')
            updateTotal.testsdonebmxrepeat = updateTotal.testsdonebmxrepeat + form.cleaned_data.get('testsdonebmxrepeat')
            updateTotal.testsdonegenexpertrun = updateTotal.testsdonegenexpertrun + form.cleaned_data.get('testsdonegenexpertrun')
            updateTotal.testsdonegenexpertfailedbuteligibaleforrepeat = updateTotal.testsdonegenexpertfailedbuteligibaleforrepeat + form.cleaned_data.get('testsdonegenexpertfailedbuteligibaleforrepeat')
            updateTotal.testsdonegenexpertfailedbutnoteligibaleforrepeat = updateTotal.testsdonegenexpertfailedbutnoteligibaleforrepeat + form.cleaned_data.get('testsdonegenexpertfailedbutnoteligibaleforrepeat')
            updateTotal.testsdonegenexpertrepeat = updateTotal.testsdonegenexpertrepeat + form.cleaned_data.get('testsdonegenexpertrepeat')


            updateTotal.testsdonequantstudio3run = updateTotal.testsdonequantstudio3run + form.cleaned_data.get('testsdonequantstudio3run')
            updateTotal.testsdonequantstudio3failedbuteligibaleforrepeat = updateTotal.testsdonequantstudio3failedbuteligibaleforrepeat + form.cleaned_data.get('testsdonequantstudio3failedbuteligibaleforrepeat')
            updateTotal.testsdonequantstudio3failedbutnoteligibaleforrepeat = updateTotal.testsdonequantstudio3failedbutnoteligibaleforrepeat + form.cleaned_data.get('testsdonequantstudio3failedbutnoteligibaleforrepeat')
            updateTotal.testsdonequantstudio3repeat = updateTotal.testsdonequantstudio3repeat + form.cleaned_data.get('testsdonequantstudio3repeat')


            updateTotal.testsdonehologicpantherrun = updateTotal.testsdonehologicpantherrun + form.cleaned_data.get('testsdonehologicpantherrun')
            updateTotal.testsdonehologicpantherfailedbuteligibaleforrepeat = updateTotal.testsdonehologicpantherfailedbuteligibaleforrepeat + form.cleaned_data.get('testsdonehologicpantherfailedbuteligibaleforrepeat')
            updateTotal.testsdonehologicpantherfailedbutnoteligibaleforrepeat = updateTotal.testsdonehologicpantherfailedbutnoteligibaleforrepeat + form.cleaned_data.get('testsdonehologicpantherfailedbutnoteligibaleforrepeat')
            updateTotal.testsdonehologicpantherrepeat = updateTotal.testsdonehologicpantherrepeat + form.cleaned_data.get('testsdonehologicpantherrepeat')
            updateTotal.testsdonerdtabrun = updateTotal.testsdonerdtabrun + form.cleaned_data.get('testsdonerdtabrun')


            updateTotal.testsdonerdtabfailedbuteligibaleforrepeat = updateTotal.testsdonerdtabfailedbuteligibaleforrepeat + form.cleaned_data.get('testsdonerdtabfailedbuteligibaleforrepeat')
            updateTotal.testsdonerdtabfailedbutnoteligibaleforrepeat = updateTotal.testsdonerdtabfailedbutnoteligibaleforrepeat + form.cleaned_data.get('testsdonerdtabfailedbutnoteligibaleforrepeat')
            updateTotal.testsdonerdtabrepeat = updateTotal.testsdonerdtabrepeat + form.cleaned_data.get('testsdonerdtabrepeat')


            updateTotal.testsdonerdtagrun = updateTotal.testsdonerdtagrun + form.cleaned_data.get('testsdonerdtagrun')
            updateTotal.testsdonerdtagfailedbuteligibaleforrepeat = updateTotal.testsdonerdtagfailedbuteligibaleforrepeat + form.cleaned_data.get('testsdonerdtagfailedbuteligibaleforrepeat')
            updateTotal.testsdonerdtagfailedbutnoteligibaleforrepeat = updateTotal.testsdonerdtagfailedbutnoteligibaleforrepeat + form.cleaned_data.get('testsdonerdtagfailedbutnoteligibaleforrepeat')
            
            updateTotal.testsdonerdtagrepeat = updateTotal.testsdonerdtagrepeat + form.cleaned_data.get('testsdonerdtagrepeat')
            updateTotal.totaltestsdone = updateTotal.totaltestsdone + form.cleaned_data.get('totaltestsdone')
            updateTotal.totalrepeats = updateTotal.totalrepeats + form.cleaned_data.get('totalrepeats')
            updateTotal.totalpatientsrun = updateTotal.totalpatientsrun + form.cleaned_data.get('totalpatientsrun')

            updateTotal.errorratesabbott = updateTotal.errorratesabbott + form.cleaned_data.get('errorratesabbott')
            updateTotal.errorratesbmx = updateTotal.errorratesbmx + form.cleaned_data.get('errorratesbmx')
            updateTotal.errorratesgenexpert = updateTotal.errorratesgenexpert + form.cleaned_data.get('errorratesgenexpert')
            updateTotal.errorratesquantstudio3 = updateTotal.errorratesquantstudio3 + form.cleaned_data.get('errorratesquantstudio3')
            updateTotal.errorrateshologicpanther = updateTotal.errorrateshologicpanther + form.cleaned_data.get('errorrateshologicpanther')
            updateTotal.errorratesrdtab = updateTotal.errorratesrdtab + form.cleaned_data.get('errorratesrdtab')
            updateTotal.errorratesrdtag = updateTotal.errorratesrdtag + form.cleaned_data.get('errorratesrdtag')

           
            # updateTotal.save()

            updateTotal.save()

            new_form = form.save(commit=False)

            new_form.lab =request.user.lab
            new_form.reportingweek = getReportingWeek()


            new_form.user=request.user
            new_form.save()


            #return render(request, 'success.html')
        else:
            print(form.errors.as_data())
            #return render(request, 'success.html')
    form = Specimensruncovid19Form()
    context = {'form': form, 'vl':False,'eid':False,'covid':True}

    return render(request, 'masvingo_brti_covid_19_weekly_statistics_tool_31-6_June_2021/Specimens_Run.html', context)

@login_required(login_url='loly_login')
def labcov19recieved(request):
    # updateTotal = Specimensreceivedcovid19.objects.get(dayofweek="Total")
    # for row in updateTotal.__dict__:
    #         print(updateTotal.__dict__[row].)

    if request.method == 'POST':
        form = Specimensreceivedcovid19Form(request.POST)
        if form.is_valid():
            print('e=nter tge rhtiung')

            updateTotal = specimensreceivedbrticovid19.objects.get(dayofweek="Total",lab='brti',reportingweek='1')
            updateTotal.samplescarriedoverpreviousweeks = updateTotal.samplescarriedoverpreviousweeks + form.cleaned_data.get('samplescarriedoverpreviousweeks')
            updateTotal.samplesreceivedcurrentweeknasopharyngealswab = updateTotal.samplesreceivedcurrentweeknasopharyngealswab + form.cleaned_data.get('samplesreceivedcurrentweeknasopharyngealswab')
            updateTotal.samplesreceivedcurrentweeknasalswab = updateTotal.samplesreceivedcurrentweeknasalswab + form.cleaned_data.get('samplesreceivedcurrentweeknasalswab')
            updateTotal.samplesreceivedcurrentweekoropharyngealswab = updateTotal.samplesreceivedcurrentweekoropharyngealswab + form.cleaned_data.get('samplesreceivedcurrentweekoropharyngealswab')

            updateTotal.samplesreceivedcurrentweekmidturbinateswab = updateTotal.samplesreceivedcurrentweekmidturbinateswab + form.cleaned_data.get('samplesreceivedcurrentweekmidturbinateswab')
            updateTotal.samplesreceivedcurrentweeksputum = updateTotal.samplesreceivedcurrentweeksputum + form.cleaned_data.get('samplesreceivedcurrentweeksputum')
            updateTotal.samplesreceivedcurrentweekwholebloodorplasmaorserum = updateTotal.samplesreceivedcurrentweekwholebloodorplasmaorserum + form.cleaned_data.get('samplesreceivedcurrentweekwholebloodorplasmaorserum')
            updateTotal.samplesreceivedcurrentweekother = updateTotal.samplesreceivedcurrentweekother + form.cleaned_data.get('samplesreceivedcurrentweekother')

            updateTotal.samplesrejectedcurrentweek = updateTotal.samplesrejectedcurrentweek + form.cleaned_data.get('samplesrejectedcurrentweek')
            updateTotal.totalsamplesreceivedcurrentweek = updateTotal.totalsamplesreceivedcurrentweek + form.cleaned_data.get('totalsamplesreceivedcurrentweek')
            updateTotal.numberofsamplesenteredintolims = updateTotal.numberofsamplesenteredintolims + form.cleaned_data.get('numberofsamplesenteredintolims')
            updateTotal.totalsamplescurrentpluscarryover = updateTotal.totalsamplescurrentpluscarryover + form.cleaned_data.get('totalsamplescurrentpluscarryover')

            updateTotal.samplesreferred = updateTotal.samplesreferred + form.cleaned_data.get('samplesreferred')
            updateTotal.rejectionratecurrentweek = updateTotal.rejectionratecurrentweek + form.cleaned_data.get('rejectionratecurrentweek')
            updateTotal.numberofresultsprintedlims = updateTotal.numberofresultsprintedlims + form.cleaned_data.get('numberofresultsprintedlims')
            updateTotal.totalresultsdispatchedbylab = updateTotal.totalresultsdispatchedbylab + form.cleaned_data.get('totalresultsdispatchedbylab')

#mostt likely to crush if we keep on joining comments 
            updateTotal.comment = updateTotal.comment + form.cleaned_data.get('comment')
            updateTotal.samplesreferredtoname = updateTotal.samplesreferredtoname + form.cleaned_data.get('samplesreferredtoname')





            updateTotal.save()

            new_form = form.save(commit=False)

            new_form.lab =request.user.lab
            new_form.reportingweek = getReportingWeek()


            new_form.user=request.user
            new_form.save()


            #return render(request, 'success.html')
        else:
            print(form.errors.as_data())
    form = Specimensreceivedcovid19Form()
    context = {'form': form , 'vl':False,'eid':False,'covid':True}
    return render(request, 'masvingo_brti_covid_19_weekly_statistics_tool_31-6_June_2021/Specimens_Received.html', context)

@login_required(login_url='loly_login')
def labcov19general(request):
    if request.method == 'POST':
        form = Generalcovid19Form(request.POST)
        if form.is_valid():
            print('form is valid',request.user.lab)
            updateTotal = generalbrticovid19.objects.get(dayofweek="Total",lab=request.user.lab,reportingweek=getReportingWeek())
            updateTotal.commentsregardingtestingandchallengesfacedbythelaboratory = updateTotal.commentsregardingtestingandchallengesfacedbythelaboratory + form.cleaned_data.get('commentsregardingtestingandchallengesfacedbythelaboratory')
            updateTotal.numberofstaffwhotestedpositivetocovid19atvllab = updateTotal.numberofstaffwhotestedpositivetocovid19atvllab + form.cleaned_data.get('numberofstaffwhotestedpositivetocovid19atvllab')
            updateTotal.numberOfStaffwhotestedpositivetocovid19athubs = updateTotal.numberOfStaffwhotestedpositivetocovid19athubs + form.cleaned_data.get('numberOfStaffwhotestedpositivetocovid19athubs')
            updateTotal.numberofstaffwhohavebeenvaccinated = updateTotal.numberofstaffwhohavebeenvaccinated + form.cleaned_data.get('numberofstaffwhohavebeenvaccinated')
            updateTotal.Requesttobrtifromthelaboratory = updateTotal.Requesttobrtifromthelaboratory + form.cleaned_data.get('Requesttobrtifromthelaboratory')
            updateTotal.Comments = updateTotal.Comments + form.cleaned_data.get('Comments')
            
            
            updateTotal.save()

            new_form = form.save(commit=False)

            new_form.lab =request.user.lab
            new_form.reportingweek = getReportingWeek()


            new_form.user=request.user
            new_form.save()


            #return render(request, 'success.html')
        else:
            print(form.errors.as_data())
    form = Generalcovid19Form()
    context = {'form': form , 'vl':False,'eid':False,'covid':True}
    return render(request, 'masvingo_brti_covid_19_weekly_statistics_tool_31-6_June_2021/general.html', context)

@login_required(login_url='loly_login')
def labcov19machine(request):
    if request.method == 'POST':
        print('tyuiop[')
        form = machinedowntimereagentstockouttoolcovid19Form(request.POST)
        if form.is_valid():
            print('tyuiop[ valid')
            updateTotal = machinedowntimereagentstockouttoolbrticovid19.objects.get(dayofweek="Total", lab=request.user.lab,reportingweek=getReportingWeek())
            updateTotal.numberofmachinebreakdownsabbott += form.cleaned_data.get('numberofmachinebreakdownsabbott')
            updateTotal.numberofmachinebreakdownsbmx = updateTotal.numberofmachinebreakdownsbmx + form.cleaned_data.get('numberofmachinebreakdownsbmx')
            updateTotal.numberofmachinebreakdownsgenexpert = updateTotal.numberofmachinebreakdownsgenexpert + form.cleaned_data.get('numberofmachinebreakdownsgenexpert')
            updateTotal.numberofmachinebreakdownsquantstudio3 = updateTotal.numberofmachinebreakdownsquantstudio3 + form.cleaned_data.get('numberofmachinebreakdownsquantstudio3')
           
            updateTotal.numberofmachinebreakdownshologicpanther = updateTotal.numberofmachinebreakdownshologicpanther + form.cleaned_data.get('numberofmachinebreakdownshologicpanther')
            updateTotal.numberofmachinebreakdownscomments = updateTotal.numberofmachinebreakdownscomments + form.cleaned_data.get('numberofmachinebreakdownscomments')
            updateTotal.machinedowntimedaysabbott = updateTotal.machinedowntimedaysabbott + form.cleaned_data.get('machinedowntimedaysabbott')
            updateTotal.machinedowntimedaysbmx = updateTotal.machinedowntimedaysbmx + form.cleaned_data.get('machinedowntimedaysbmx')
           
            updateTotal.machinedowntimedaysgenexpert = updateTotal.machinedowntimedaysgenexpert + form.cleaned_data.get('machinedowntimedaysgenexpert')
            updateTotal.machinedowntimedaysquantstudio3 = updateTotal.machinedowntimedaysquantstudio3 + form.cleaned_data.get('machinedowntimedaysquantstudio3')
            updateTotal.machinedowntimedayshologicpanther = updateTotal.machinedowntimedayshologicpanther + form.cleaned_data.get('machinedowntimedayshologicpanther')
            updateTotal.machinedowntimedayscomments = updateTotal.machinedowntimedayscomments + form.cleaned_data.get('machinedowntimedayscomments')
           
            updateTotal.reagentstockoutabbort = updateTotal.reagentstockoutabbort + form.cleaned_data.get('reagentstockoutabbort')
            updateTotal.reagentstockoutbms = updateTotal.reagentstockoutbms + form.cleaned_data.get('reagentstockoutbms')
            updateTotal.reagentstockoutgenexpert = updateTotal.reagentstockoutgenexpert + form.cleaned_data.get('reagentstockoutgenexpert')
            
            updateTotal.reagentstockoutquantstudio3 = updateTotal.reagentstockoutquantstudio3 + form.cleaned_data.get('reagentstockoutquantstudio3')
            updateTotal.reagentstockouthologicpanther = updateTotal.reagentstockouthologicpanther + form.cleaned_data.get('reagentstockouthologicpanther')
            updateTotal.reagentstockoutcomments = updateTotal.reagentstockoutcomments + form.cleaned_data.get('reagentstockoutcomments')
            updateTotal.reagentstockoutcomments = updateTotal.reagentstockoutcomments + form.cleaned_data.get('reagentstockoutcomments')

            # updateTotal.Requesttobrtifromthelaboratory = updateTotal.Requesttobrtifromthelaboratory + form.cleaned_data.get('Requesttobrtifromthelaboratory
            # updateTotal.numberofstaffwhohavebeenvaccinated = updateTotal.numberofstaffwhohavebeenvaccinated + form.cleaned_data.get('numberofstaffwhohavebeenvaccinated
            updateTotal.save()

            new_form = form.save(commit=False)

            new_form.lab =request.user.lab
            new_form.reportingweek = getReportingWeek()


            new_form.user=request.user
            new_form.save()


            #return render(request, 'success.html')
        else:
            print(form.errors.as_data())
            #return render(request, 'success.html')
    form = machinedowntimereagentstockouttoolcovid19Form()
    context = {'form': form , 'vl':False,'eid':False,'covid':True}
    return render(request, 'masvingo_brti_covid_19_weekly_statistics_tool_31-6_June_2021/Machine_Downtime_&_Reagent_stock_out_tool.html', context)





#-------------------------------------------lab eid-----------------------------------------------------------------------------------------

def labeidtop1(request):
    if request.method == 'POST':
        form = TopbrtiweeklyForm(request.POST)
        if form.is_valid():
            form.save()
            #return render(request, 'success.html')
    form = TopbrtiweeklyForm()
    context = {'form': form , 'vl':True,'eid':False,'covid':False}
    return render(request, 'masvingo_brti_weekly_statistics_tool_june_2021/top.html', context)

def labeidtop2(request):
    if request.method == 'POST':
        form = TopbrtiweeklyForm(request.POST)
        if form.is_valid():
            form.save()
            #return render(request, 'success.html')
    form = TopbrtiweeklyForm()
    context = {'form': form , 'vl':True,'eid':False,'covid':False}
    return render(request, 'masvingo_brti_weekly_statistics_tool_june_2021/top.html', context)




def labeidtop3(request):
    if request.method == 'POST':
        form = TopbrtiweeklyForm(request.POST)
        if form.is_valid():
            form.save()
            #return render(request, 'success.html')
    form = TopbrtiweeklyForm()
    context = {'form': form , 'vl':True,'eid':False,'covid':False}
    return render(request, 'masvingo_brti_weekly_statistics_tool_june_2021/top.html', context)

def labeidtop4(request):
    if request.method == 'POST':
        form = TopbrtiweeklyForm(request.POST)
        if form.is_valid():
            form.save()
            #return render(request, 'success.html')
    form = TopbrtiweeklyForm()
    context = {'form': form , 'vl':True,'eid':False,'covid':False}
    return render(request, 'masvingo_brti_weekly_statistics_tool_june_2021/top.html', context)

def labeidtop5(request):
    if request.method == 'POST':
        form = TopbrtiweeklyForm(request.POST)
        if form.is_valid():
            form.save()
            #return render(request, 'success.html')
    form = TopbrtiweeklyForm()
    context = {'form': form , 'vl':True,'eid':False,'covid':False}
    return render(request, 'masvingo_brti_weekly_statistics_tool_june_2021/top.html', context)


@login_required(login_url='loly_login')
def labeidtop6(request):
    if request.method == 'POST':
        form = TopbrtiweeklyForm(request.POST)
        if form.is_valid():
            form.save()
            #return render(request, 'success.html')
    form = TopbrtiweeklyForm()
    context = {'form': form , 'vl':True,'eid':False,'covid':False}
    return render(request, 'masvingo_brti_weekly_statistics_tool_june_2021/top.html', context)


def labeidtop7(request):
    if request.method == 'POST':
        form = TopbrtiweeklyForm(request.POST)
        if form.is_valid():
            form.save()
            #return render(request, 'success.html')
    form = TopbrtiweeklyForm()
    context = {'form': form , 'vl':True,'eid':False,'covid':False}
    return render(request, 'masvingo_brti_weekly_statistics_tool_june_2021/top.html', context)








def labeidrun(request):
    if request.method == 'POST':
        form = SpecimensrunbrtiweeklyForm(request.POST)
        if form.is_valid():

            updateTotal = specimensrunbrtivleid.objects.get(key="eid",dayofweek="Total", lab=request.user.lab,reportingweek=getReportingWeek())
            updateTotal.testsdonerochenumberofsamplesreceivedthisweekplasma += form.cleaned_data.get('testsdonerochenumberofsamplesreceivedthisweekplasma')
            updateTotal.testsdonerochenumberofsamplescarriedoverpreviousweeksplasma += form.cleaned_data.get('testsdonerochenumberofsamplescarriedoverpreviousweeksplasma')
            updateTotal.testsdonerochefailedbuteligibaleforrepeatplasma += form.cleaned_data.get('testsdonerochefailedbuteligibaleforrepeatplasma')
            updateTotal.testsdonerochefailedbutnoteligibaleforrepeatplasma += form.cleaned_data.get('testsdonerochefailedbutnoteligibaleforrepeatplasma')


            updateTotal.testsdonerocherepeatplasma += form.cleaned_data.get('testsdonerocherepeatplasma')
            updateTotal.testsdonerochefailedafterrepeattestingplasma += form.cleaned_data.get('testsdonerochefailedafterrepeattestingplasma')
            updateTotal.testsdonerochenumberofsamplesreceivedthisweekdbs += form.cleaned_data.get('testsdonerochenumberofsamplesreceivedthisweekdbs')
            updateTotal.testsdonerochenumberofsamplescarriedoverpreviousweeksdbs += form.cleaned_data.get('testsdonerochenumberofsamplescarriedoverpreviousweeksdbs')
            updateTotal.testsdonerochefailedbuteligibaleforrepeatdbs += form.cleaned_data.get('testsdonerochefailedbuteligibaleforrepeatdbs')


            updateTotal.testsdonerochefailedbutnoteligibaleforrepeatdbs += form.cleaned_data.get('testsdonerochefailedbutnoteligibaleforrepeatdbs')
            updateTotal.testsdonerocherepeatdbs += form.cleaned_data.get('testsdonerocherepeatdbs')
            updateTotal.testsdonerochefailedafterrepeattestingdbs += form.cleaned_data.get('testsdonerochefailedafterrepeattestingdbs')
            updateTotal.testsdonebmxnumberofsamplesreceivedthisweekplasma += form.cleaned_data.get('testsdonebmxnumberofsamplesreceivedthisweekplasma')
            updateTotal.testsdonebmxnumberofsamplescarriedoverpreviousweeksplasma += form.cleaned_data.get('testsdonebmxnumberofsamplescarriedoverpreviousweeksplasma')

            updateTotal.testsdonebmxfailedbuteligibaleforrepeatplasma += form.cleaned_data.get('testsdonebmxfailedbuteligibaleforrepeatplasma')
            updateTotal.testsdonebmxfailedbutnoteligibaleforrepeatplasma += form.cleaned_data.get('testsdonebmxfailedbutnoteligibaleforrepeatplasma')
            updateTotal.testsdonebmxrepeatplasma += form.cleaned_data.get('testsdonebmxrepeatplasma')
            updateTotal.testsdonebmxfailedafterrepeattestingplasma += form.cleaned_data.get('testsdonebmxfailedafterrepeattestingplasma')
            updateTotal.testsdonebmxnumberofsamplesreceivedthisweekdbs += form.cleaned_data.get('testsdonebmxnumberofsamplesreceivedthisweekdbs')
            updateTotal.testsdonebmxnumberofsamplescarriedoverpreviousweeksdbs += form.cleaned_data.get('testsdonebmxnumberofsamplescarriedoverpreviousweeksdbs')
            updateTotal.testsdonebmxfailedbuteligibaleforrepeatdbs += form.cleaned_data.get('testsdonebmxfailedbuteligibaleforrepeatdbs')

            updateTotal.testsdonebmxfailedbutnoteligibaleforrepeatdbs += form.cleaned_data.get('testsdonebmxfailedbutnoteligibaleforrepeatdbs')
            updateTotal.testsdonebmxrepeatdbs += form.cleaned_data.get('testsdonebmxrepeatdbs')
            updateTotal.testsdonebmxfailedafterrepeattestingdbs += form.cleaned_data.get('testsdonebmxfailedafterrepeattestingdbs')
            updateTotal.testsdoneabbottnumberofsamplesreceivedthisweekplasma += form.cleaned_data.get('testsdoneabbottnumberofsamplesreceivedthisweekplasma')
            updateTotal.testsdoneabbottnumberofsamplescarriedoverpreviousweeksplasma += form.cleaned_data.get('testsdoneabbottnumberofsamplescarriedoverpreviousweeksplasma')
            updateTotal.testsdoneabbottfailedbuteligibaleforrepeatplasma += form.cleaned_data.get('testsdoneabbottfailedbuteligibaleforrepeatplasma')
            updateTotal.testsdoneabbottfailedbutnoteligibaleforrepeatplasma += form.cleaned_data.get('testsdoneabbottfailedbutnoteligibaleforrepeatplasma')

            updateTotal.testsdoneabbottrepeatplasma += form.cleaned_data.get('testsdoneabbottrepeatplasma')
            updateTotal.testsdoneabbottfailedafterrepeattestingplasma += form.cleaned_data.get('testsdoneabbottfailedafterrepeattestingplasma')
            updateTotal.testsdoneabbottnumberofsamplesreceivedthisweekdbs += form.cleaned_data.get('testsdoneabbottnumberofsamplesreceivedthisweekdbs')
            updateTotal.testsdoneabbottnumberofsamplescarriedoverpreviousweeksdbs += form.cleaned_data.get('testsdoneabbottnumberofsamplescarriedoverpreviousweeksdbs')
            updateTotal.testsdoneabbottfailedbuteligibaleforrepeatdbs += form.cleaned_data.get('testsdoneabbottfailedbuteligibaleforrepeatdbs')
            updateTotal.testsdoneabbottfailedbutnoteligibaleforrepeatdbs += form.cleaned_data.get('testsdoneabbottfailedbutnoteligibaleforrepeatdbs')
            updateTotal.testsdoneabbottrepeatdbs += form.cleaned_data.get('testsdoneabbottrepeatdbs')

            updateTotal.testsdoneabbottfailedafterrepeattestingdbs += form.cleaned_data.get('testsdoneabbottfailedafterrepeattestingdbs')
            updateTotal.testsdonehologicpanthernumberofsamplesreceivedthisweekplasma += form.cleaned_data.get('testsdonehologicpanthernumberofsamplesreceivedthisweekplasma')
            updateTotal.tests2 += form.cleaned_data.get('tests2')
            updateTotal.testsdonehologicpantherfailedbuteligibaleforrepeatplasma += form.cleaned_data.get('testsdonehologicpantherfailedbuteligibaleforrepeatplasma')
            updateTotal.testsdonehologicpantherfailedbutnoteligibaleforrepeatplasma += form.cleaned_data.get('testsdonehologicpantherfailedbutnoteligibaleforrepeatplasma')
            updateTotal.testsdonehologicpantherrepeatplasma += form.cleaned_data.get('testsdonehologicpantherrepeatplasma')
            updateTotal.testsdonehologicpantherfailedafterrepeattestingplasma += form.cleaned_data.get('testsdonehologicpantherfailedafterrepeattestingplasma')

            updateTotal.testsdonehologicpanthernumberofsamplesreceivedthisweekdbs += form.cleaned_data.get('testsdonehologicpanthernumberofsamplesreceivedthisweekdbs')
            updateTotal.tests1 += form.cleaned_data.get('tests1')
            updateTotal.testsdonehologicpantherfailedbuteligibaleforrepeatdbs += form.cleaned_data.get('testsdonehologicpantherfailedbuteligibaleforrepeatdbs')
            updateTotal.testsdonehologicpantherfailedbutnoteligibaleforrepeatdbs += form.cleaned_data.get('testsdonehologicpantherfailedbutnoteligibaleforrepeatdbs')
            updateTotal.testsdonehologicpantherrepeatdbs += form.cleaned_data.get('testsdonehologicpantherrepeatdbs')
            updateTotal.testsdonehologicpantherfailedafterrepeattestingdbs += form.cleaned_data.get('testsdonehologicpantherfailedafterrepeattestingdbs')

            updateTotal.totaltestsdone += form.cleaned_data.get('totaltestsdone')
            updateTotal.totalrepeats += form.cleaned_data.get('totalrepeats')
            updateTotal.totalpatientsrun += form.cleaned_data.get('totalpatientsrun')
            updateTotal.targetsweekly += form.cleaned_data.get('targetsweekly')
            updateTotal.percentagetargetsachievements += form.cleaned_data.get('percentagetargetsachievements')
            updateTotal.percentageerrorraterocheplasma += form.cleaned_data.get('percentageerrorraterocheplasma')
            updateTotal.percentageerrorraterochedbs += form.cleaned_data.get('percentageerrorraterochedbs')

            updateTotal.percentageerrorratebmxplasma += form.cleaned_data.get('percentageerrorratebmxplasma')
            updateTotal.percentageerrorratebmxdbs += form.cleaned_data.get('percentageerrorratebmxdbs')
            updateTotal.percentageerrorrateabbottplasma += form.cleaned_data.get('percentageerrorrateabbottplasma')
            updateTotal.percentageerrorrateabbottdbs += form.cleaned_data.get('percentageerrorrateabbottdbs')
            updateTotal.percentageerrorratehologicpantherplasma += form.cleaned_data.get('percentageerrorratehologicpantherplasma')
            updateTotal.percentageerrorratehologicpantherdbs += form.cleaned_data.get('percentageerrorratehologicpantherdbs')
            updateTotal.totalncsfromaudit += form.cleaned_data.get('totalncsfromaudit')
            updateTotal.ncsnotyetclosed += form.cleaned_data.get('ncsnotyetclosed')

            updateTotal.ncsclosedthisweek += form.cleaned_data.get('ncsclosedthisweek')

            
            updateTotal.save()

            new_form = form.save(commit=False)

            new_form.lab =request.user.lab
            new_form.reportingweek = getReportingWeek()


            new_form.user=request.user
            new_form.save()


            #return render(request, 'success.html')
        else:
            print(form.errors.as_data())
            #return render(request, 'success.html')
    form = SpecimensrunbrtiweeklyForm()
    context = {'form': form, 'vl':False,'eid':True,'covid':False}

    return render(request, 'masvingo_brti_weekly_statistics_tool_june_2021/Specimensrun.html', context)

@login_required(login_url='loly_login')
def labeidrecieved(request):
    if request.method == 'POST':
        form = SpecimensreceivedbrtiweeklyForm(request.POST)
        if form.is_valid():

            updateTotal = specimensreceivedbrtivleid.objects.get(key="eid",dayofweek="Total", lab=request.user.lab,reportingweek=getReportingWeek())
            updateTotal.samplescarriedoverpreviousweeksnevertestedplasma += form.cleaned_data.get('samplescarriedoverpreviousweeksnevertestedplasma')
            updateTotal.samplescarriedoverpreviousweeknevertesteddbs += form.cleaned_data.get('samplescarriedoverpreviousweeksnevertesteddbs')
            updateTotal.samplescarriedoverpreviousfailedsamplesplasma += form.cleaned_data.get('samplescarriedoverpreviousfailedsamplesplasma')
            updateTotal.samplescarriedoverpreviousfailedsamplesdbs += form.cleaned_data.get('samplescarriedoverpreviousfailedsamplesdbs')

            updateTotal.samplesreceivedcurrentweekplasma += form.cleaned_data.get('samplesreceivedcurrentweekplasma')
            updateTotal.samplesreceivedcurrentweekdbs += form.cleaned_data.get('samplesreceivedcurrentweekdbs')
            updateTotal.samplesrejectedcurrentweekplasma += form.cleaned_data.get('samplesrejectedcurrentweekplasma')
            updateTotal.samplesrejectedcurrentweekdbs += form.cleaned_data.get('samplesrejectedcurrentweekdbs')

            updateTotal.totalsamplesreceivedcurrentweekplasma += form.cleaned_data.get('totalsamplesreceivedcurrentweekplasma')
            updateTotal.totalsamplesreceivedcurrentweekdbs += form.cleaned_data.get('totalsamplesreceivedcurrentweekdbs')
            updateTotal.numberofsamplesenteredintolimsondayofarrivalplasma += form.cleaned_data.get('numberofsamplesenteredintolimsondayofarrivalplasma')
            updateTotal.numberofsamplesenteredintolimsondayofarrivaldbs += form.cleaned_data.get('numberofsamplesenteredintolimsondayofarrivaldbs')
            updateTotal.numberofsamplesenteredintolimsafterdayofarrivalplasma += form.cleaned_data.get('numberofsamplesenteredintolimsafterdayofarrivalplasma')

            updateTotal.numberofsamplesenteredintolimsafterdayofarrivaldbs += form.cleaned_data.get('numberofsamplesenteredintolimsafterdayofarrivaldbs')
            updateTotal.numberofhourslimswasfunctional += form.cleaned_data.get('numberofhourslimswasfunctional')
            updateTotal.totalsamplescurrentandcarryoverplasma += form.cleaned_data.get('totalsamplescurrentandcarryoverplasma')
            updateTotal.totalsamplescurrentandcarryoverdbs += form.cleaned_data.get('totalsamplescurrentandcarryoverdbs')
            updateTotal.samplesrefferedplasma += form.cleaned_data.get('samplesrefferedplasma')

            updateTotal.samplesreffereddbs += form.cleaned_data.get('samplesreffereddbs')
            updateTotal.labsamplesrefferedto += form.cleaned_data.get('labsamplesrefferedto')
            updateTotal.percentagerejectionrateplasma += form.cleaned_data.get('percentagerejectionrateplasma')
            updateTotal.percentagerejectionratedbs += form.cleaned_data.get('percentagerejectionratedbs')
            updateTotal.numberofresultsprintedfromlimsplasma += form.cleaned_data.get('numberofresultsprintedfromlimsplasma')
            updateTotal.numberofresultsprintedfromlimsdbs += form.cleaned_data.get('numberofresultsprintedfromlimsdbs')

            updateTotal.totalresultsdispatchedbylabplasma += form.cleaned_data.get('totalresultsdispatchedbylabplasma')
            updateTotal.totalresultsdispatchedbylabdbs += form.cleaned_data.get('totalresultsdispatchedbylabdbs')
            updateTotal.totalresultsdispatchedbylabviasmsplasma += form.cleaned_data.get('totalresultsdispatchedbylabviasmsplasma')
            updateTotal.totalresultsdispatchedbylabviasmsdbs += form.cleaned_data.get('totalresultsdispatchedbylabviasmsdbs')
            updateTotal.reasonsforrejectionssamplequalitycompromisedplasma += form.cleaned_data.get('reasonsforrejectionssamplequalitycompromisedplasma')

            updateTotal.reasonsforrejectionssamplequalitycompromiseddbs += form.cleaned_data.get('reasonsforrejectionssamplequalitycompromiseddbs')
            updateTotal.reasons6 += form.cleaned_data.get('reasons6')
            updateTotal.reasons5 += form.cleaned_data.get('reasons5')
            updateTotal.reasons2 += form.cleaned_data.get('reasons2')
            updateTotal.reasons1 += form.cleaned_data.get('reasons1')

            updateTotal.reasons4 += form.cleaned_data.get('reasons4')
            updateTotal.reason3 += form.cleaned_data.get('reasons3')
            updateTotal.reasonsforrejectionssamplequalitycompromisedqdacheckplasma += form.cleaned_data.get('reasonsforrejectionssamplequalitycompromisedqdacheckplasma')
            updateTotal.reasonsforrejectionssamplequalitycompromisedqdacheckdbs += form.cleaned_data.get('reasonsforrejectionssamplequalitycompromisedqdacheckdbs')
            updateTotal.reasonsforsamplerefferalreagentorkitstockoutplasma += form.cleaned_data.get('reasonsforsamplerefferalreagentorkitstockoutplasma')

            updateTotal.reasonsforsamplerefferalreagentorkitstockoutdbs += form.cleaned_data.get('reasonsforsamplerefferalreagentorkitstockoutdbs')
            updateTotal.reasonsforsamplerefferalinstrumentmechanicalfailureplasma += form.cleaned_data.get('reasonsforsamplerefferalinstrumentmechanicalfailureplasma')
            updateTotal.reasonsforsamplerefferalinstrumentmechanicalfailuredbs += form.cleaned_data.get('reasonsforsamplerefferalinstrumentmechanicalfailuredbs')
            updateTotal.reasonsforsamplerefferalinsufficientinstrumentcapacityplasma += form.cleaned_data.get('reasonsforsamplerefferalinsufficientinstrumentcapacityplasma')
            updateTotal.reasonsforsamplerefferalinsufficientinstrumentcapacitydbs += form.cleaned_data.get('reasonsforsamplerefferalinsufficientinstrumentcapacitydbs')

            updateTotal.reasonsforsamplerefferalinsufficienthrcapacityplasma += form.cleaned_data.get('reasonsforsamplerefferalinsufficienthrcapacityplasma')
            updateTotal.reasonsforsamplerefferalinsufficienthrcapacitydbs += form.cleaned_data.get('reasonsforsamplerefferalinsufficienthrcapacitydbs')
            updateTotal.reasonsforsamplerefferaldqacheckplasma += form.cleaned_data.get('reasonsforsamplerefferaldqacheckplasma')
            updateTotal.reasonsforsamplerefferaldqacheckdbs += form.cleaned_data.get('reasonsforsamplerefferaldqacheckdbs')
            updateTotal.comments += form.cleaned_data.get('comments')

            updateTotal.save()

            new_form = form.save(commit=False)

            new_form.lab =request.user.lab
            new_form.reportingweek = getReportingWeek()


            new_form.user=request.user
            new_form.save()


            #return render(request, 'success.html')
        else:
            print(form.errors.as_data())
            #return render(request, 'success.html')
    form = SpecimensreceivedbrtiweeklyForm()
    context = {'form': form , 'vl':False,'eid':True,'covid':False}
    return render(request, 'masvingo_brti_weekly_statistics_tool_june_2021/specimenreceived.html', context)

@login_required(login_url='loly_login')
def labeidfailure(request):
    if request.method == 'POST':
        form = ReasonsforfailurebrtiweeklyForm(request.POST)
        if form.is_valid():
            updateTotal = reasonsforfailurebrtivleid.objects.get(key="eid",dayofweek="Total", lab=request.user.lab,reportingweek=getReportingWeek())
            updateTotal.rocheplasmanumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetosamplequalityissues')
            updateTotal.rocheplasmanumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetoreagentqualityissues')
            updateTotal.rocheplasmanumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetoduetoqcfailure')
            updateTotal.rocheplasmanumberoffailedtestsduetopowerfailure += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetopowerfailure')
            updateTotal.rocheplasmanumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetomechanicalfailure')
            updateTotal.rocheplasmanumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetoprocessingerror')
            updateTotal.rocheplasmanumberoffailedtestsduetoother += form.cleaned_data.get('rocheplasmanumberoffailedtestsduetoother')



            updateTotal.rochedqacheckplasma += form.cleaned_data.get('rochedqacheckplasma')
            updateTotal.rochedbsnumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('rochedbsnumberoffailedtestsduetosamplequalityissues')
            updateTotal.rochedbsnumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('rochedbsnumberoffailedtestsduetoreagentqualityissues')
            updateTotal.rochedbsnumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('rochedbsnumberoffailedtestsduetoduetoqcfailure')
            updateTotal.rochedbsnumberoffailedtestsduetopowerfailure += form.cleaned_data.get('rochedbsnumberoffailedtestsduetopowerfailure')
            updateTotal.rochedbsnumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('rochedbsnumberoffailedtestsduetomechanicalfailure')
            updateTotal.rochedbsnumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('rochedbsnumberoffailedtestsduetoprocessingerror')

            updateTotal.rochedbsnumberoffailedtestsduetoother += form.cleaned_data.get('rochedbsnumberoffailedtestsduetoother')
            updateTotal.rochedqacheckdbs += form.cleaned_data.get('rochedqacheckdbs')

            updateTotal.bmxplasmanumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetosamplequalityissues')
            updateTotal.bmxplasmanumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetoreagentqualityissues')
            updateTotal.bmxplasmanumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetoduetoqcfailure')

            updateTotal.bmxplasmanumberoffailedtestsduetopowerfailure += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetopowerfailure')
            updateTotal.bmxplasmanumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetomechanicalfailure')
            updateTotal.bmxplasmanumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetoprocessingerror')
            updateTotal.bmxplasmanumberoffailedtestsduetoother += form.cleaned_data.get('bmxplasmanumberoffailedtestsduetoother')
            updateTotal.bmxdqacheckplasma += form.cleaned_data.get('bmxdqacheckplasma')
            updateTotal.bmxdbsnumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetosamplequalityissues')

            updateTotal.bmxdbsnumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetoreagentqualityissues')
            updateTotal.bmxdbsnumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetoduetoqcfailure')
            updateTotal.bmxdbsnumberoffailedtestsduetopowerfailure += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetopowerfailure')
            updateTotal.bmxdbsnumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetomechanicalfailure')
            updateTotal.bmxdbsnumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetoprocessingerror')
            updateTotal.bmxdbsnumberoffailedtestsduetoother += form.cleaned_data.get('bmxdbsnumberoffailedtestsduetoother')
            updateTotal.bmxdqacheckdbs += form.cleaned_data.get('bmxdqacheckdbs')

            updateTotal.abbottplasmanumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetosamplequalityissues')
            updateTotal.abbottplasmanumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetoreagentqualityissues')
            updateTotal.abbottplasmanumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetoduetoqcfailure')
            updateTotal.abbottplasmanumberoffailedtestsduetopowerfailure += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetopowerfailure')
            updateTotal.abbottplasmanumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetomechanicalfailure')
            updateTotal.abbottplasmanumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetoprocessingerror')
            updateTotal.abbottplasmanumberoffailedtestsduetoother += form.cleaned_data.get('abbottplasmanumberoffailedtestsduetoother')

            updateTotal.abbottdqacheckplasma += form.cleaned_data.get('abbottdqacheckplasma')
            updateTotal.abbottdbsnumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetosamplequalityissues')
            updateTotal.abbottdbsnumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetoreagentqualityissues')
            updateTotal.abbottdbsnumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetoduetoqcfailure')
            updateTotal.abbottdbsnumberoffailedtestsduetopowerfailure += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetopowerfailure')
            updateTotal.abbottdbsnumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetomechanicalfailure')
            updateTotal.abbottdbsnumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetoprocessingerror')



            updateTotal.abbottdbsnumberoffailedtestsduetoother += form.cleaned_data.get('abbottdbsnumberoffailedtestsduetoother')
            updateTotal.abbottdqacheckdbs += form.cleaned_data.get('abbottdqacheckdbs')
            updateTotal.hologicpantherplasmanumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('hologicpantherplasmanumberoffailedtestsduetosamplequalityissues')
            updateTotal.hpantherplasmanumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('hpantherplasmanumberoffailedtestsduetoreagentqualityissues')
            updateTotal.hologicpantherplasmanumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('hologicpantherplasmanumberoffailedtestsduetoduetoqcfailure')



            updateTotal.hologicpantherplasmanumberoffailedtestsduetopowerfailure += form.cleaned_data.get('hologicpantherplasmanumberoffailedtestsduetopowerfailure')
            updateTotal.hologicpantherplasmanumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('hologicpantherplasmanumberoffailedtestsduetomechanicalfailure')
            updateTotal.hologicpantherplasmanumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('hologicpantherplasmanumberoffailedtestsduetoprocessingerror')
            updateTotal.hologicpantherplasmanumberoffailedtestsduetoother += form.cleaned_data.get('hologicpantherplasmanumberoffailedtestsduetoother')
            updateTotal.hologicpantherdqacheckplasma += form.cleaned_data.get('hologicpantherdqacheckplasma')
            updateTotal.hologicpantherdbsnumberoffailedtestsduetosamplequalityissues += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetosamplequalityissues')

            updateTotal.hologicpantherdbsnumberoffailedtestsduetoreagentqualityissues += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetoreagentqualityissues')
            updateTotal.hologicpantherdbsnumberoffailedtestsduetoduetoqcfailure += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetoduetoqcfailure')
            updateTotal.hologicpantherdbsnumberoffailedtestsduetopowerfailure += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetopowerfailure')
            updateTotal.hologicpantherdbsnumberoffailedtestsduetomechanicalfailure += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetomechanicalfailure')
            updateTotal.hologicpantherdbsnumberoffailedtestsduetoprocessingerror += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetoprocessingerror')
            updateTotal.hologicpantherdbsnumberoffailedtestsduetoother += form.cleaned_data.get('hologicpantherdbsnumberoffailedtestsduetoother')
            updateTotal.hologicpantherdqacheckdbs += form.cleaned_data.get('hologicpantherdqacheckdbs')

           
            updateTotal.save()

            new_form = form.save(commit=False)

            new_form.lab =request.user.lab
            new_form.reportingweek = getReportingWeek()


            new_form.user=request.user
            new_form.save()


            #return render(request, 'success.html')
        else:
            print(form.errors.as_data())
            #return render(request, 'success.html')
    form = ReasonsforfailurebrtiweeklyForm()
    context = {'form': form , 'vl':False,'eid':True,'covid':False}
    return render(request, 'masvingo_brti_weekly_statistics_tool_june_2021/Reasonsforfailures.html', context)

@login_required(login_url='loly_login')
def labeidelectric(request):
    if request.method == 'POST':
        form = ElectricoutagebrtiweeklyForm(request.POST)
        if form.is_valid():


            updateTotal = Electricoutagebrtivleid.objects.get(key="eid",dayofweek="Monday", lab=request.user.lab,reportingweek=getReportingWeek())
            updateTotal.numberofhourswithnoelectricityperday += form.cleaned_data.get('numberofhourswithnoelectricityperday')
            updateTotal.numberofhoursgeneratorwasonperday += form.cleaned_data.get('numberofhoursgeneratorwasonperday')
            updateTotal.litresoffueladdedtogeneratorperday += form.cleaned_data.get('litresoffueladdedtogeneratorperday')
            updateTotal.numberofhoursmachineswasnotbeingusedduetopowercutperday += form.cleaned_data.get('numberofhoursmachineswasnotbeingusedduetopowercutperday')
            updateTotal.totaltestsdoneperdayusinggenerator += form.cleaned_data.get('totaltestsdoneperdayusinggenerator')
           
            updateTotal.save()

            new_form = form.save(commit=False)

            new_form.lab =request.user.lab
            new_form.reportingweek = getReportingWeek()


            new_form.user=request.user
            new_form.save()


            #return render(request, 'success.html')
        else:
            print(form.errors.as_data())
            #return render(request, 'success.html')
    form = ElectricoutagebrtiweeklyForm()
    context = {'form': form , 'vl':False,'eid':True,'covid':False}
    return render(request, 'masvingo_brti_weekly_statistics_tool_june_2021/electricoutagestool.html', context)


from .panjabi import *
from openpyxl import *

def populateVleidSheets(labs):
    workbook = load_workbook(filename="vleid.xlsx")
    #lst = workbook.sheetnames


    # ws = workbook[workbook.sheetnames[int(reporting_week)]]
    # sheet = workbook.active
    #generalcovidlst=getCovidGeneralList()

    vleidsreceivedlist =getVlSpecimenRunList()
    for lab in labs:
        for week in ['1','2','3','4']:
            ws = workbook[workbook.sheetnames[int(week)-1]]
            sheet = workbook.active

            getRecord(week,lab,'vl','specimenrun',ws= sheet,lst=getVlSpecimenRunList())
            getRecord(week,lab,'eid','specimenrun',ws= sheet,lst=getVlSpecimenRunList())

            getRecord(week,lab,'vl','specimenreceived',ws= sheet,lst=vleidsreceivedlist)
            getRecord(week,lab,'eid','specimenreceived',ws= sheet,lst=vleidsreceivedlist)


            getRecord(week,lab,'vl','reasons',ws= sheet,lst=getCovidMachineList())
            getRecord(week,lab,'eid','reasons',ws= sheet,lst=getCovidMachineList())

            getRecord(week,lab,'vl','machine',ws= sheet,lst=getCovidGeneralList())
            # getRecord(week,lab,'covid','general2',ws= sheet,lst=generalcovidlst)
            # getRecord(week,lab,'covid','general3',ws= sheet,lst=generalcovidlst)



    workbook.save(filename="hello_eid.xlsx")



def populateCovidSheets(labs):
    workbook = load_workbook(filename="covid.xlsx")
    #lst = workbook.sheetnames


    # ws = workbook[workbook.sheetnames[int(reporting_week)]]
    # sheet = workbook.active
    generalcovidlst=getCovidGeneralList()

    for lab in labs:
        for week in ['1','2','3','4']:
            ws = workbook[workbook.sheetnames[int(week)-1]]
            sheet = workbook.active
            getRecord(week,lab,'covid','specimenrun',ws= sheet,lst=getCovidSpecimenRunList())

            getRecord(week,lab,'covid','specimenreceived',ws= sheet,lst=getCovidSpecimenReceiveList())
            getRecord(week,lab,'covid','machine',ws= sheet,lst=getCovidMachineList())
            getRecord(week,lab,'covid','general1',ws= sheet,lst=generalcovidlst)
            getRecord(week,lab,'covid','general2',ws= sheet,lst=generalcovidlst)
            getRecord(week,lab,'covid','general3',ws= sheet,lst=generalcovidlst)



    workbook.save(filename="hello_world.xlsx")






@login_required(login_url='loly_login')
def export_csv(request):
 


    #covid.xlsx
    labscovid=[
        'NMRL',
        'NTBRL',
        'Mutare',
        'Wilkins',
        'BRIDH',
        'Gweru',
        'Chinhoyi',
        'Masvingo',
        'Victoria Falls', 
        'Bindura',
        'Kadoma',
        'Marondera',
        'St Lukes',
        'Gwanda']

    populateCovidSheets(labscovid)

    # populateVleidSheets(labsvleid)


    from openpyxl.writer.excel import save_virtual_workbook
    workbook = load_workbook(filename="hello_world.xlsx")
    data =save_virtual_workbook(workbook)



    response=HttpResponse(content=data,content_type='application/ms-excel')
    response['Content-Disposition']='attachment; filename=covid' + '.xlsx'




    # workbook = load_workbook(filename="C:\\Users\\kaste\\OneDrive\\Desktop\\meals postgres\\meal\\pages\\covid.xlsx")
    # #lst = workbook.sheetnames
    # ws = workbook[workbook.sheetnames[0]]
    # sheet = workbook.active


    # getRecord(getReportingWeek(),request.user.lab,'covid','specimenrun',workbook=workbook ,ws= sheet,lst=getCovidSpecimenRunList())

    # getRecord(getReportingWeek(),request.user.lab,'covid','specimenreceived',workbook=workbook ,ws= sheet,lst=getCovidSpecimenReceiveList())
    # getRecord(getReportingWeek(),request.user.lab,'covid','machine',workbook=workbook ,ws= sheet,lst=getCovidMachineList())
    # getRecord(getReportingWeek(),request.user.lab,'covid','general1',workbook=workbook ,ws= sheet,lst=getCovidGeneralList())
    # getRecord(getReportingWeek(),request.user.lab,'covid','general2',workbook=workbook ,ws= sheet,lst=getCovidGeneralList())
    # getRecord(getReportingWeek(),request.user.lab,'covid','general3',workbook=workbook ,ws= sheet,lst=getCovidGeneralList())

    return response
    #return covidlist


#   1st lab nmrl positions from 0 to 17
#   2st lab nmrl positions from 18 to 35


def getRecord(reporting_week,lab,stats,table,ws,lst):


    
    if stats =='covid':
        if table == 'specimenrun':
            start = 0
            end = 0
            res = getPositions(lab,stats,table)
            try:
                record = specimensrunbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
                pos=0
                for key in record.__dict__:
                    if key =='_state':
                        pass
                    elif key =='id':
                        pass
                    elif key == 'dayofweek':
                        pass
                    elif key =='reportingweek':
                        pass
                    elif key == 'lab':
                        pass
                    else:
                        try:
                            ws[lst[res][pos]]=record.__dict__[key]
                            #print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                            pos +=1
                        except IndexError as e:
                            print(e)
                            break
            except:
                print('record for lab: '+lab+' week: '+reporting_week+'not found')
        elif table == 'specimenreceived':
            start = 0
            end = 0           
            res = getPositions(lab,stats,table)
            try:
                record = specimensreceivedbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
                #usr_names = [str(elem) for elem in list(record.values_list())]

                #print( print (dir(record))
                pos=0
                for key in record.__dict__:
                    if key =='_state':
                        pass
                    elif key =='id':
                        pass
                    elif key == 'dayofweek':
                        pass
                    elif key =='reportingweek':
                        pass
                    elif key == 'lab':
                        pass
                    else:
                        try:
                            ws[lst[res][pos]]=record.__dict__[key]
                            print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                            pos +=1
                        except IndexError as e:
                            print(e)
                            break
                
            except:
                print('record for lab: '+lab+' week: '+reporting_week+'not found')



        elif table == 'general1':
            start = 0
            end = 0           
            res = getPositions(lab,stats,'general1')
            try:
                record = generalbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
                #usr_names = [str(elem) for elem in list(record.values_list())]
                print(record.__dict__)

                lst1=lst[0:14]

                #print( print (dir(record))
                pos=0
                for key in record.__dict__:
                    if key =='_state':
                        pass
                    elif key =='id':
                        pass
                    elif key == 'dayofweek':
                        pass
                    elif key =='reportingweek':
                        pass
                    elif key == 'lab':
                        pass

                    elif key=='commentsoninteruptions':
                        try:
                            ws[lst1[res][pos]]=record.__dict__[key]
                            print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                            pos +=1
                            break
                        except IndexError as e:
                            print(e)
                            break

            except:
                print('record for lab: '+lab+' week: '+reporting_week+'not found')
        elif table == 'general2':
            start = 0
            end = 0           
            res = getPositions(lab,stats,'general2') 
            try:
                record = generalbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
                #usr_names = [str(elem) for elem in list(record.values_list())]
                lst1=lst[15:29]
                #print( print (dir(record))
                pos=0
                for key in record.__dict__:
                    if key =='_state':
                        pass
                    elif key =='id':
                        pass
                    elif key == 'dayofweek':
                        pass
                    elif key =='reportingweek':
                        pass
                    elif key == 'lab':
                        pass
                    elif key=='numberofstaffwhotestedpositivetocovid19atvllab':
                        try:
                            ws[lst1[res][0]]=record.__dict__[key]
                            print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                            pos +=1
                        except IndexError as e:
                            print(e)
                    elif key=='numberOfStaffwhotestedpositivetocovid19athubs':
                        try:
                            ws[lst1[res][1]]=record.__dict__[key]
                            print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                            pos +=1
                        except IndexError as e:
                            print(e)
                    elif key=='numberofstaffwhohavebeenvaccinated':
                        try:
                            ws[lst1[res][2]]=record.__dict__[key]
                            print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                            pos +=1
                            
                        except IndexError as e:
                            print(e)
                    elif key=='Comments':
                        try:
                            ws[lst1[res][3]]=record.__dict__[key]
                            print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                            pos +=1
                            break
                        except IndexError as e:
                            print(e)
            except:
                print('record for lab: '+lab+' week: '+reporting_week+'not found')                            
        elif table == 'general3':
            start = 0
            end = 0           
            res = getPositions(lab,stats,'general3')
            lst1=lst[30:]
            try:
                record = generalbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
                #usr_names = [str(elem) for elem in list(record.values_list())]

                #print( print (dir(record))
                pos=0
                for key in record.__dict__:
                    if key =='_state':
                        pass
                    elif key =='id':
                        pass
                    elif key == 'dayofweek':
                        pass
                    elif key =='reportingweek':
                        pass
                    elif key == 'lab':
                        pass
                    elif key =='Requesttobrtifromthelaboratory':
                        try:
                            ws[lst1[10][0]]=record.__dict__[key]
                            print('assigning '+str(record.__dict__[key]),'to '+str(lst1[res][pos])+' index '+str(pos))
                            pos +=1
                            break
                        except IndexError as e:
                            print(e)
                            break
            except:
                print('record for lab: '+lab+' week: '+reporting_week+'not found')

        elif table == 'machine':
            start = 0
            end = 0           
            res = getPositions(lab,stats,table)
            try:
                record = machinedowntimereagentstockouttoolbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
                #usr_names = [str(elem) for elem in list(record.values_list())]

                #print( print (dir(record))
                pos=0
                for key in record.__dict__:
                    if key =='_state':
                        pass
                    elif key =='id':
                        pass
                    elif key == 'dayofweek':
                        pass
                    elif key =='reportingweek':
                        pass
                    elif key == 'lab':
                        pass
                    else:
                        try:
                            ws[lst[res][pos]]=record.__dict__[key]
                            print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                            pos +=1
                        except IndexError as e:
                            print(e)
                            break
            except:
                print('record for lab: '+lab+' week: '+reporting_week+'not found')

    # if stats =='vl':
    #     if table == 'specimenrun':
    #         start = 0
    #         end = 0
    #         res = getPositions(lab,stats,table)
    #         record = specimensrunbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
    #         pos=0
    #         for key in record.__dict__:
    #             if key =='_state':
    #                 pass
    #             elif key =='id':
    #                 pass
    #             elif key == 'dayofweek':
    #                 pass
    #             elif key =='reportingweek':
    #                 pass
    #             elif key == 'lab':
    #                 pass
    #             else:
    #                 try:
    #                     ws[lst[res][pos]]=record.__dict__[key]
    #                     print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
    #                     pos +=1
    #                 except IndexError as e:
    #                     print(e)
    #     elif table == 'specimenreceived':
    #         start = 0
    #         end = 0           
    #         res = getPositions(lab,stats,table)

    #         record = specimensreceivedbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
    #         #usr_names = [str(elem) for elem in list(record.values_list())]

    #         #print( print (dir(record))
    #         pos=0
    #         for key in record.__dict__:
    #             if key =='_state':
    #                 pass
    #             elif key =='id':
    #                 pass
    #             elif key == 'dayofweek':
    #                 pass
    #             elif key =='reportingweek':
    #                 pass
    #             elif key == 'lab':
    #                 pass
    #             else:
    #                 ws[lst[pos+res[0]]]=record.__dict__[key]
    #                 print('assigning '+str(record.__dict__[key]),'to '+str(lst[pos+res[0]]))
    #                 pos +=1

    #     elif table == 'general':
    #         start = 0
    #         end = 0           
    #         res = getPositions(lab,stats,table)

    #         record = specimensreceivedbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
    #         #usr_names = [str(elem) for elem in list(record.values_list())]

    #         #print( print (dir(record))
    #         pos=0
    #         for key in record.__dict__:
    #             if key =='_state':
    #                 pass
    #             elif key =='id':
    #                 pass
    #             elif key == 'dayofweek':
    #                 pass
    #             elif key =='reportingweek':
    #                 pass
    #             elif key == 'lab':
    #                 pass
    #             else:
    #                 ws[lst[pos+res]]=record.__dict__[key]
    #                 #print('assigning '+str(record.__dict__[key]),'to '+str(lst[pos+res[0]]))
    #                 pos +=1

    #     elif table == 'machine':
    #         start = 0
    #         end = 0           
    #         res = getPositions(lab,stats,table)

    #         record = specimensreceivedbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
    #         #usr_names = [str(elem) for elem in list(record.values_list())]

    #         #print( print (dir(record))
    #         pos=0
    #         for key in record.__dict__:
    #             if key =='_state':
    #                 pass
    #             elif key =='id':
    #                 pass
    #             elif key == 'dayofweek':
    #                 pass
    #             elif key =='reportingweek':
    #                 pass
    #             elif key == 'lab':
    #                 pass
    #             else:
    #                 ws[lst[pos+res]]=record.__dict__[key]
    #                 print('assigning '+str(record.__dict__[key]),'to '+str(lst[pos+res[0]]))
    #                 pos +=1





#***************************vleid******************************************************************************************************


    if stats =='vl':
        if table == 'specimenrun':
                    start = 0
                    end = 0
                    res = getPositions1(lab,stats,table)
                    if res == -1:
                        return 0
                    try:
                        record = specimensrunbrtivleid.objects.get(lab=lab,reportingweek=reporting_week, key='vl')
                        print(record.__dict__)
                        pos=0
                        for key in record.__dict__:
                            if key =='_state':
                                pass
                            elif key =='id':
                                pass
                            elif key == 'dayofweek':
                                pass
                            elif key =='reportingweek':
                                pass
                            elif key == 'lab':
                                pass
                            else:
                                try:
                                    ws[lst[res][pos]]=record.__dict__[key]
                                    #print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                                    pos +=1
                                except IndexError as e:
                                    print(e)
                                    break
                    except:
                        print('record for lab: '+lab+' week: '+reporting_week+'not found')
        elif table == 'specimenreceived':
                    start = 0
                    end = 0           
                    res = getPositions(lab,stats,table)
                    try:
                        record = specimensreceivedbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
                        #usr_names = [str(elem) for elem in list(record.values_list())]

                        #print( print (dir(record))
                        pos=0
                        for key in record.__dict__:
                            if key =='_state':
                                pass
                            elif key =='id':
                                pass
                            elif key == 'dayofweek':
                                pass
                            elif key =='reportingweek':
                                pass
                            elif key == 'lab':
                                pass
                            else:
                                try:
                                    ws[lst[res][pos]]=record.__dict__[key]
                                    print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                                    pos +=1
                                except IndexError as e:
                                    print(e)
                                    break
                        
                    except:
                        print('record for lab: '+lab+' week: '+reporting_week+'not found')



        elif table == 'general1':
                    start = 0
                    end = 0           
                    res = getPositions(lab,stats,'general1')
                    try:
                        record = generalbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
                        #usr_names = [str(elem) for elem in list(record.values_list())]
                        print(record.__dict__)

                        lst1=lst[0:14]

                        #print( print (dir(record))
                        pos=0
                        for key in record.__dict__:
                            if key =='_state':
                                pass
                            elif key =='id':
                                pass
                            elif key == 'dayofweek':
                                pass
                            elif key =='reportingweek':
                                pass
                            elif key == 'lab':
                                pass

                            elif key=='commentsoninteruptions':
                                try:
                                    ws[lst1[res][pos]]=record.__dict__[key]
                                    print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                                    pos +=1
                                    break
                                except IndexError as e:
                                    print(e)
                                    break

                    except:
                        print('record for lab: '+lab+' week: '+reporting_week+'not found')
        elif table == 'general2':
                    start = 0
                    end = 0           
                    res = getPositions(lab,stats,'general2') 
                    try:
                        record = generalbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
                        #usr_names = [str(elem) for elem in list(record.values_list())]
                        lst1=lst[15:29]
                        #print( print (dir(record))
                        pos=0
                        for key in record.__dict__:
                            if key =='_state':
                                pass
                            elif key =='id':
                                pass
                            elif key == 'dayofweek':
                                pass
                            elif key =='reportingweek':
                                pass
                            elif key == 'lab':
                                pass
                            elif key=='numberofstaffwhotestedpositivetocovid19atvllab':
                                try:
                                    ws[lst1[res][0]]=record.__dict__[key]
                                    print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                                    pos +=1
                                except IndexError as e:
                                    print(e)
                            elif key=='numberOfStaffwhotestedpositivetocovid19athubs':
                                try:
                                    ws[lst1[res][1]]=record.__dict__[key]
                                    print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                                    pos +=1
                                except IndexError as e:
                                    print(e)
                            elif key=='numberofstaffwhohavebeenvaccinated':
                                try:
                                    ws[lst1[res][2]]=record.__dict__[key]
                                    print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                                    pos +=1
                                    
                                except IndexError as e:
                                    print(e)
                            elif key=='Comments':
                                try:
                                    ws[lst1[res][3]]=record.__dict__[key]
                                    print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                                    pos +=1
                                    break
                                except IndexError as e:
                                    print(e)
                    except:
                        print('record for lab: '+lab+' week: '+reporting_week+'not found')                            
        elif table == 'general3':
                    start = 0
                    end = 0           
                    res = getPositions(lab,stats,'general3')
                    lst1=lst[30:]
                    try:
                        record = generalbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
                        #usr_names = [str(elem) for elem in list(record.values_list())]

                        #print( print (dir(record))
                        pos=0
                        for key in record.__dict__:
                            if key =='_state':
                                pass
                            elif key =='id':
                                pass
                            elif key == 'dayofweek':
                                pass
                            elif key =='reportingweek':
                                pass
                            elif key == 'lab':
                                pass
                            elif key =='Requesttobrtifromthelaboratory':
                                try:
                                    ws[lst1[10][0]]=record.__dict__[key]
                                    print('assigning '+str(record.__dict__[key]),'to '+str(lst1[res][pos])+' index '+str(pos))
                                    pos +=1
                                    break
                                except IndexError as e:
                                    print(e)
                                    break
                    except:
                        print('record for lab: '+lab+' week: '+reporting_week+'not found')

        elif table == 'machine':
                    start = 0
                    end = 0           
                    res = getPositions(lab,stats,table)
                    try:
                        record = machinedowntimereagentstockouttoolbrticovid19.objects.get(lab=lab,reportingweek=reporting_week)
                        #usr_names = [str(elem) for elem in list(record.values_list())]

                        #print( print (dir(record))
                        pos=0
                        for key in record.__dict__:
                            if key =='_state':
                                pass
                            elif key =='id':
                                pass
                            elif key == 'dayofweek':
                                pass
                            elif key =='reportingweek':
                                pass
                            elif key == 'lab':
                                pass
                            else:
                                try:
                                    ws[lst[res][pos]]=record.__dict__[key]
                                    print('assigning '+str(record.__dict__[key]),'to '+str(lst[res][pos])+' index '+str(pos))
                                    pos +=1
                                except IndexError as e:
                                    print(e)
                                    break
                    except:
                        print('record for lab: '+lab+' week: '+reporting_week+'not found')


























            



def getPositions(lab,stats,table):
       
            if lab == 'NMRL' :
                return 0

            if lab == 'NTBRL' :
                return 1

            if lab == 'Mutare' :
                return 2

            if lab == 'Wilkins' :
                return 3
            if lab == 'BRIDH'  :
                return 4
            if lab == 'Gweru'  :
                return 5
            if lab == 'Chinhoyi'  :
                return 6
            if lab == 'Masvingo'  :
                return 7

            if lab == 'Beitbridge'  :
                return 8

            if lab == 'Victoria Falls'  :
                return 9

            if lab == 'Bindura'  :
                return 10


            if lab == 'Kadoma'  :
                return 11
            if lab == 'Marondera'  :
                return 11

            if lab == 'St Lukes'  :
                return 13

            if lab == 'Gwanda'  :
                return 14




def getPositions1(lab,stats,table):
            if lab == 'NMRL' :
                return 0
       
            if lab == 'Mpilo' :
                return 1

            if lab == 'Mutare' :
                return 2

            if lab == 'BRIDH' :
                return 3

            if lab == 'Gweru'  :
                return 4
            if lab == 'Chinhoyi'  :
                return 5
            if lab == 'Masvingo'  :
                return 6

            if lab == 'Victoria Falls'  :
                return 7

            if lab == 'Bindura'  :
                return 8


            if lab == 'Kadoma'  :
                return 9
            if lab == 'Marondera'  :
                return 10

            if lab == 'St Lukes'  :
                return 11

            if lab == 'Gwanda'  :
                return 12

            return -1

def getPositions2(lab,stats,table):
            if lab == 'NMRL' :
                return 13
       
            if lab == 'Mpilo' :
                return 14

            if lab == 'Mutare' :
                return 15


            if lab == 'Chinhoyi'  :
                return 16
            if lab == 'Masvingo'  :
                return 17
            return -1

  





